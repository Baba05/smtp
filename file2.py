import os
import re
from openpyxl import load_workbook
#block to read data from excel file and appending to list "real[]"
myexcel=load_workbook(r'C:\Users\Admin\Desktop\data.xlsx')          
totalsheet=myexcel.sheetnames
mysheet=myexcel[totalsheet[0]]
rows=mysheet.max_row
cols=mysheet.max_column
print(rows,cols)
real=[]
startpos=1

while(rows>0):
    instr='A'+str(startpos)
    appstr='B'+str(startpos)
    datestr='C'+str(startpos)
    errstr='D'+str(startpos)
    statusstr='E'+str(startpos)
    incident=mysheet[instr].value
    application=mysheet[appstr].value
    date=mysheet[datestr].value
    error=mysheet[errstr].value
    status=mysheet[statusstr].value
    real.append([incident,application,date,error,status])
    rows-=1
    startpos+=1
#print(" list",real)

# block to searching issues and appending to particular list
l1=[]#network issue
l2=[]#database
l3=[]#Acess denied
for x in real:
    if re.findall('network',x[3]):
        l1.append(x)
    elif re.findall('databse',x[3]):
        l2.append(x)
    elif re.findall('Access',x[3]):
        l3.append(x)

'''print("NETWORK",l1)
print("@@@@@@@@@@@@@@@@@")
print("DATABASE",l2)
print("@@@@@@@@@@@@@@@@@")
print("ACCESS DENIED",l3)'''

# block to insert list l1 l2 l3 to txt file
with open(r'C:\Users\Admin\Desktop\bob.txt', 'w') as filehandle:
    data="NETWORK ISSUE---------------------------\n"
    filehandle.write(data)
    for listitem in l1:
        filehandle.write('%s\n' % listitem)
    data1="DATABASE ISSUE---------------------------\n"
    filehandle.write(data1)
    for listitem in l2:
        filehandle.write('%s\n' % listitem)
    data2="ACCESS DENIED ISSUE---------------------------\n"
    filehandle.write(data2)
    for listitem in l3:
        filehandle.write('%s\n' % listitem)
#-----------------------------------------------------------------------------#
#sending mail
import smtplib
from email.mime.multipart import MIMEMultipart  #importing
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

frm = "@@@@@@@@@@@@@gmail.com"            #reciver and sender
to = "########@gmail.com"

msg = MIMEMultipart()           #calling 

msg['From'] = frm
msg['To'] = to
msg['Subject'] = "SUBJECT OF THE EMAIL"

body = "TEXT YOU WANT TO SEND" #message to show

msg.attach(MIMEText(body, 'plain'))

filename = "bob.txt"
attachment = open(r"C:\Users\Admin\Desktop\bob.txt",'rb') #declaring name and location of file
print("1")

part = MIMEBase('application', 'octet-stream')
part.set_payload((attachment).read())
encoders.encode_base64(part)
print("2")
part.add_header('Content-Disposition', 'attachment', filename= filename)

msg.attach(part)

server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(frm, "############")            #chcking password
text = msg.as_string()
server.sendmail(frm, to, text)
server.quit()
#_--------------------------------------------------------------------------#


    
    

